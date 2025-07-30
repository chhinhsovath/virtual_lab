import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create timeline data
timeline_data = [
    # August 2024
    {
        "Month": "August 2024",
        "Week": "Week 1-2",
        "Dates": "Aug 1-14",
        "Activity": "System Setup & Deployment",
        "Description": "Deploy Virtual Lab LMS to production, configure database with TaRL tables, set up 65 teacher accounts across 32 schools",
        "Deliverables": "Production system deployed, user accounts created, permissions configured",
        "Number of Engineers": 3,
        "Person-Days": 30,
        "Responsible": "Technical Team"
    },
    {
        "Month": "August 2024",
        "Week": "Week 1-2",
        "Dates": "Aug 1-14",
        "Activity": "Database Configuration",
        "Description": "Configure school access permissions, subject assignments, integrate with existing TaRL tables",
        "Deliverables": "Database fully configured with access controls",
        "Number of Engineers": 2,
        "Person-Days": 20,
        "Responsible": "Database Team"
    },
    {
        "Month": "August 2024",
        "Week": "Week 3-4",
        "Dates": "Aug 15-31",
        "Activity": "Teacher Training & Onboarding",
        "Description": "Conduct training for 65 teachers on system usage, bilingual interface, student selection, assessment entry",
        "Deliverables": "All teachers trained, user guides created, practice sessions completed",
        "Number of Engineers": 2,
        "Person-Days": 20,
        "Responsible": "Training Team + Tech Support"
    },
    {
        "Month": "August 2024",
        "Week": "Week 3-4",
        "Dates": "Aug 15-31",
        "Activity": "Documentation & Support Materials",
        "Description": "Create user guides in Khmer and English, video tutorials, FAQ documents",
        "Deliverables": "Complete documentation package in both languages",
        "Number of Engineers": 1,
        "Person-Days": 10,
        "Responsible": "Documentation Team"
    },
    
    # September 2024
    {
        "Month": "September 2024",
        "Week": "Week 1",
        "Dates": "Sep 2-6",
        "Activity": "Student Selection Phase",
        "Description": "Teachers select 20 students per school using drag-and-drop interface, total 640 students",
        "Deliverables": "All TaRL students selected and registered in system",
        "Number of Engineers": 1,
        "Person-Days": 5,
        "Responsible": "Tech Support"
    },
    {
        "Month": "September 2024",
        "Week": "Week 2-3",
        "Dates": "Sep 9-20",
        "Activity": "Baseline Assessment Execution",
        "Description": "Conduct baseline assessments for Khmer and Math, data entry for all 640 students",
        "Deliverables": "Complete baseline data for all students",
        "Number of Engineers": 2,
        "Person-Days": 20,
        "Responsible": "Tech Support + Data Team"
    },
    {
        "Month": "September 2024",
        "Week": "Week 4",
        "Dates": "Sep 23-27",
        "Activity": "Baseline Data Analysis",
        "Description": "Generate reports, analyze student levels, prepare intervention strategies",
        "Deliverables": "Baseline reports, student level analysis, intervention plans",
        "Number of Engineers": 2,
        "Person-Days": 10,
        "Responsible": "Data Analysis Team"
    },
    
    # October 2024
    {
        "Month": "October 2024",
        "Week": "Week 1-2",
        "Dates": "Oct 1-11",
        "Activity": "TaRL Teaching Support",
        "Description": "Technical support for teachers implementing TaRL methodology, system monitoring",
        "Deliverables": "Continuous system support, usage reports",
        "Number of Engineers": 1,
        "Person-Days": 10,
        "Responsible": "Tech Support"
    },
    {
        "Month": "October 2024",
        "Week": "Week 3-4",
        "Dates": "Oct 14-25",
        "Activity": "Midline Assessment Implementation",
        "Description": "Support midline assessments, ensure data quality, generate progress reports",
        "Deliverables": "Complete midline data, progress tracking reports",
        "Number of Engineers": 2,
        "Person-Days": 20,
        "Responsible": "Tech Support + Data Team"
    },
    {
        "Month": "October 2024",
        "Week": "Week 5",
        "Dates": "Oct 28-31",
        "Activity": "Mid-Program Analysis",
        "Description": "Analyze baseline vs midline data, generate comparison reports, identify trends",
        "Deliverables": "Comparative analysis reports, improvement metrics",
        "Number of Engineers": 2,
        "Person-Days": 8,
        "Responsible": "Data Analysis Team"
    },
    
    # November 2024
    {
        "Month": "November 2024",
        "Week": "Week 1-2",
        "Dates": "Nov 4-15",
        "Activity": "Endline Assessment Support",
        "Description": "Final assessment round support, ensure complete data collection for all cycles",
        "Deliverables": "Complete endline data for all students",
        "Number of Engineers": 2,
        "Person-Days": 20,
        "Responsible": "Tech Support + Data Team"
    },
    {
        "Month": "November 2024",
        "Week": "Week 3",
        "Dates": "Nov 18-22",
        "Activity": "Comprehensive Reporting",
        "Description": "Generate final reports showing full progression cycle, create data exports for provincial office",
        "Deliverables": "Complete impact reports, data exports, analytics dashboard",
        "Number of Engineers": 3,
        "Person-Days": 15,
        "Responsible": "Data Analysis + Reporting Team"
    },
    {
        "Month": "November 2024",
        "Week": "Week 4",
        "Dates": "Nov 25-29",
        "Activity": "Program Evaluation & Planning",
        "Description": "Evaluate system performance, collect feedback, plan improvements for next cycle",
        "Deliverables": "Evaluation report, improvement recommendations, next cycle plan",
        "Number of Engineers": 2,
        "Person-Days": 10,
        "Responsible": "Project Management Team"
    }
]

# Create DataFrame
df = pd.DataFrame(timeline_data)

# Calculate totals
total_person_days = df['Person-Days'].sum()
avg_engineers = df['Number of Engineers'].mean()

# Add summary row
summary_row = pd.DataFrame([{
    "Month": "TOTAL",
    "Week": "",
    "Dates": "Aug-Nov 2024",
    "Activity": "Project Summary",
    "Description": f"Total effort across all activities",
    "Deliverables": "Complete TaRL assessment system implementation",
    "Number of Engineers": f"{avg_engineers:.1f} (avg)",
    "Person-Days": total_person_days,
    "Responsible": "All Teams"
}])

df = pd.concat([df, summary_row], ignore_index=True)

# Create Excel file
output_file = '/Users/user/Desktop/apps/virtual_lab/docs/Virtual_Lab_Timeline_August_November_2024.xlsx'
writer = pd.ExcelWriter(output_file, engine='openpyxl')
df.to_excel(writer, sheet_name='Project Timeline', index=False)

# Get the workbook and worksheet
workbook = writer.book
worksheet = writer.sheets['Project Timeline']

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=12)
summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
summary_font = Font(bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply header formatting
for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Apply cell formatting and borders
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Format summary row
        if cell.row == worksheet.max_row:
            cell.fill = summary_fill
            cell.font = summary_font

# Adjust column widths
column_widths = {
    'A': 15,  # Month
    'B': 12,  # Week
    'C': 15,  # Dates
    'D': 30,  # Activity
    'E': 50,  # Description
    'F': 40,  # Deliverables
    'G': 18,  # Number of Engineers
    'H': 15,  # Person-Days
    'I': 25   # Responsible
}

for col, width in column_widths.items():
    worksheet.column_dimensions[col].width = width

# Set row heights
worksheet.row_dimensions[1].height = 30
for row in range(2, worksheet.max_row + 1):
    worksheet.row_dimensions[row].height = 50

# Add project information at the top
worksheet.insert_rows(1, 3)
worksheet['A1'] = 'Virtual Lab LMS - TaRL Assessment System'
worksheet['A2'] = 'Project Timeline: August - November 2024'
worksheet['A3'] = f'Total Person-Days: {total_person_days} | Average Team Size: {avg_engineers:.1f} Engineers'

# Merge cells for title
worksheet.merge_cells('A1:I1')
worksheet.merge_cells('A2:I2')
worksheet.merge_cells('A3:I3')

# Format title cells
title_font = Font(size=16, bold=True)
subtitle_font = Font(size=14, bold=True)
info_font = Font(size=12, italic=True)

worksheet['A1'].font = title_font
worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
worksheet['A2'].font = subtitle_font
worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
worksheet['A3'].font = info_font
worksheet['A3'].alignment = Alignment(horizontal='center', vertical='center')

# Save the file
writer.close()

print(f"Excel file created successfully: {output_file}")
print(f"Total Person-Days: {total_person_days}")
print(f"Average Number of Engineers: {avg_engineers:.1f}")